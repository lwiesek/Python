import openpyxl
from openpyxl.styles import Font


doilu=int(input("Podaj do ilu chcesz tabliczke mnozenia: "))

wb=openpyxl.Workbook()
sheet=wb.active

for i in range(1,doilu+1):
    sheet.cell(i+1,1).value=i
    sheet.cell(i+1,1).font=Font(bold=True)
    sheet.cell(1,i+1).value=i
    sheet.cell(1,i+1).font=Font(bold=True)
    for j in range(1,doilu+1):
        sheet.cell(i+1,j+1).value=i*j

wb.save('excelTabliczka.xlsx')
wb.close()




