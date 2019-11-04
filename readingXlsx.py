import openpyxl

wb=openpyxl.load_workbook('warzywa.xlsx')
sheet=wb.active

for i in range(1,sheet.max_column+1):
    file=open("pliczek" + str(i) +".txt","w+")
    for j in range(1,sheet.max_row+1):        
        file.write(str(sheet.cell(j,i).value) + "\n")

file.close()
wb.close()
        
