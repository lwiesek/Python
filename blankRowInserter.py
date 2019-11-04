import openpyxl

N=int(input("Podaj w którym wierszu chcesz wstawic puste wiersze: "))
M=int(input("Podaj ile chcesz wstawic pustych wierszy: "))

wb1=openpyxl.load_workbook('produceSales.xlsx')
sheet1=wb1.active

sheet1.insert_rows(N,M)

wb2=openpyxl.Workbook()
sheet2=wb2.active

for j in range(1,sheet1.max_column+1):
    for i in range(1,N):
        sheet2.cell(i,j).value=sheet1.cell(i,j).value
    for i in range(N+M,sheet1.max_row+1):
        sheet2.cell(i,j).value=sheet1.cell(i,j).value

print("Kopiowanie zakonczone!")
wb2.save("produceSaleswiersze.xlsx")
wb1.close()
wb2.close()
        
