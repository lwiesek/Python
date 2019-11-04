import openpyxl

wb1=openpyxl.load_workbook('warzywa.xlsx')
sheet1=wb1.active

wb2=openpyxl.Workbook()
sheet2=wb2.active


SheetData=[]

for i in range(1,sheet1.max_row+1):
    SheetData.append([])
    for j in range(1,sheet1.max_column+1):        
        SheetData[i-1].append(sheet1.cell(i,j).value)


for i in range(0,len(SheetData)):
    for j in range(0,len(SheetData[i])):
        sheet2.cell(j+1,i+1).value=SheetData[i][j]


print("Transpozycja zakonczona!")
wb2.save("warzywaTransposed.xlsx")
wb1.close()
wb2.close()
                         
